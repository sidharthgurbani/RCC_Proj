from sklearn.model_selection import StratifiedKFold, cross_val_score
from sklearn.ensemble import RandomForestClassifier
from sklearn.base import ClassifierMixin
from sklearn.tree import DecisionTreeClassifier
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import accuracy_score
from sklearn.feature_selection import RFECV
from sklearn.base import BaseEstimator
from openpyxl import load_workbook
import xgboost as xgb
from statistics import mean
from openpyxl.styles import Font
import pandas as pd
import numpy as np
import copy

class PearsonCorr:
    df = None
    dataset = None
    new_df = None
    target = None
    X = None
    y = None
    feature_list = None
    cor = None

    def __init__(self, df, dataset, target):
        self.df = df
        self.dataset = dataset
        self.target = target

    def getPearsonCorrelation(self):
        print("Pearson Correlation!!")
        df = self.df[1:]
        self.cor = df.corr()

    def filterHighlyCorrelatedFeatures(self):
        cor = self.cor.abs()
        cols = np.full((cor.shape[0],), True, dtype=bool)
        for i in range(1, cor.shape[0]):
            for j in range(i + 1, cor.shape[0]):
                if cor.iloc[i, j] >= 0.85:
                    if cor.iloc[1, i] >= cor.iloc[1, j]:
                        cols[j] = False
                    else:
                        cols[i] = False

        selected_cols = self.df.columns[cols]
        self.new_df = self.df[selected_cols]
        self.X = self.new_df.drop(["Case", self.target], axis=1)
        self.y = self.new_df[self.target]
        self.feature_list = self.new_df.columns[2:]

    def transformed(self):
        return self.X, self.y, self.feature_list

    def saveCorrelationMatrix(self):
        name = "Dataset/pearsonCorr_" + self.dataset + ".xlsx"
        self.cor.to_excel(name)

    def saveCorrelatedFeatures(self):
        self.new_df.to_excel("Dataset/FinalData.xlsx")

class CorrMatrix(BaseEstimator):

    X_transform = None
    y_transform = None

    def __init__(self):
        super(CorrMatrix, self).__init__()

    def fit(self, X, y):
        #print("This is CorrMatrix model fit")
        y_t = np.resize(y, (y.shape[0], 1))
        data = np.hstack((y_t, X))
        df_data = pd.DataFrame(data=data)
        cor_matr = np.corrcoef(x=data, rowvar=False)
        cor = pd.DataFrame(data=cor_matr)
        cols = np.full((cor.shape[0],), True, dtype=bool)
        for i in range(1, cor.shape[0]):
            for j in range(i + 1, cor.shape[0]):
                if cor.iloc[i, j] >= 0.85:
                    if cor.iloc[0, i] >= cor.iloc[0, j]:
                        cols[j] = False
                    else:
                        cols[i] = False

        selected_cols = df_data.columns[cols]
        new_df = df_data[selected_cols]
        final_data = new_df.to_numpy()
        self.X_transform = final_data[:,1:]
        self.y_transform = final_data[:,0]
        return self

    def predict(self, X):
        return self.y_transform

    def transform(self):
        return self.X_transform, self.y_transform

class XGBC(xgb.XGBClassifier):
    model = xgb.XGBClassifier()
    feature_importance = None
    feature_list = None
    X_transformed = None

    def __init__(self, feature_list, objective="binary_:logistic"):
        super(XGBC, self).__init__(objective=objective)
        self.feature_list = feature_list

    def fit(self, X, y):
        scaler = StandardScaler()
        X = scaler.fit_transform(X)
        self.model.fit(X, y)
        self.feature_importance = self.model.feature_importances_
        self.updateList(X)
        self.model.fit(self.X_transformed, y)
        return self

    def predict(self, y):
        return self.model.predict(y)

    def score(self, X, y):
        return accuracy_score(y, self.predict(X), sample_weight=sample_weight)

    def transform(self, X):
        X_tr = self.updateList(X)
        return X_tr

    def updateList(self, X):
        indices = []
        for i, val in enumerate(self.feature_importance):
            if val > 0.01:
                indices.append(i)
        self.X_transformed = X[:, indices]


class RF(DecisionTreeClassifier):
    model = None
    feature_importance = None
    feature_list = None
    X_transformed = None
    score = None

    def __init__(self, feature_list):
        super(RF, self).__init__()
        self.feature_list = feature_list
        self.model = RandomForestClassifier(n_estimators=100, max_depth=50)

    def fit(self, X, y, sample_weight=None, check_input=True, X_idx_sorted=None):
        #print("This is RF model fit")
        scaler = StandardScaler()
        X = scaler.fit_transform(X)
        # corr = CorrMatrix()
        # corr.fit(X,y)
        # Xp, yp = corr.transform()
        # print("Final shape of data is: {}".format(Xp.shape))
        # self.model.fit(Xp,yp)
        self.model.fit(X, y)
        self.feature_importance = self.model.feature_importances_
        self.X_transformed, _ = self.updateList(X)
        self.model.fit(self.X_transformed, y)
        return self

    def predict(self, X):
        scaler = StandardScaler()
        X = scaler.fit_transform(X)
        X_tr, _ = self.updateList(X)
        return self.model.predict(X_tr)

    def score(self, X, y, sample_weight=None):
        return accuracy_score(y, self.predict(X), sample_weight=sample_weight)

    def transform(self, X):
        X_tr, _ = self.updateList(X)
        return X_tr

    def saveFeatures(self):
        filename = 'Dataset/temp.xlsx'
        wb = load_workbook(filename)
        ws = wb.create_sheet('feature_weights')
        wcell1 = ws.cell(1, 1, 'Feature')
        wcell1.font = Font(bold=True)
        wcell2 = ws.cell(1, 2, 'Weights')
        wcell2.font = Font(bold=True)
        index = 0
        for i, val in enumerate(self.feature_importance):
            if val >= 0:
                wcell1 = ws.cell(index + 2, 1, self.feature_list[i])
                wcell2 = ws.cell(index + 2, 2, val)
                index += 1

        wb.save(filename)

    def updateList(self, X):
        new_feat = list()
        indices = []
        for i, val in enumerate(self.feature_importance):
            if val > 0.01:
                new_feat.append(val)
                indices.append(i)

        X_new = X[:,indices]
        return X_new, new_feat


class nestedRFECV(DecisionTreeClassifier):
    models = []
    X_minmax = None
    clf = None
    feature_list = None
    outer_loop = 1
    inner_loop = 1
    list = None
    ranking = None
    X_transformed = None
    scores = None
    loop_indices = []

    def __init__(self, feature_list):
        self.feature_list = feature_list
        # Choose which classifier you need to use to perform RFECV with
        # self.clf = RandomForestClassifier(n_estimators=10, max_depth=20)
        self.clf = xgb.XGBClassifier()

    def fit(self, X, y, sample_weight=None, check_input=True, X_idx_sorted=None):
        # Store the original feature list and normalize the data
        list_temp = self.feature_list
        scaler = StandardScaler()
        X_minmax = scaler.fit_transform(X)
        self.X_minmax = copy.deepcopy(X_minmax)
        self.scores = []

        # Determine the number of folds to be used.
        kfold = StratifiedKFold(n_splits=5, shuffle=True)

        for outer in range(self.outer_loop):
            print("\n--------This is outer loop {}---------\n".format(outer + 1))
            # Run the outer loop from here
            for i, (train_o, test_o) in enumerate(kfold.split(X_minmax, y)):
                self.loop_indices.append((train_o, test_o))
                print("This is set {}".format(i + 1))
                X_train_o = X_minmax[train_o]
                y_train_o = y[train_o]
                X_test_o = X_minmax[test_o]
                y_test_o = y[test_o]
                X_train_transformed = copy.deepcopy(X_train_o)
                X_test_transformed = copy.deepcopy(X_test_o)

                # Run the inner loop from here
                for inner in range(self.inner_loop):
                    # If the number of features are very high (>100), we set the minimum number of features needed to be 100.
                    # If the numnber of features are moderate (15-100), we set the minimum number of features to be 10
                    # less than already present
                    n_feat = min(100, X_train_transformed.shape[1] - 10)

                    # If the number of features are less (<15), then we want it to select atleast 5 features to continue the loop
                    n_feat = max(10, n_feat)
                    list_temp_prev = list_temp
                    print("\n\t--------This is inner loop {}---------\n".format(inner + 1))
                    rfecv = RFECV(estimator=self.clf, step=1, min_features_to_select=n_feat, cv=kfold, scoring='accuracy')
                    # rfecv = xgb.XGBClassifier()

                    # Transform the datasets at each loop to keep track of reduced features
                    # rfecv.fit(X_train_transformed, y_train_o)
                    # X_train_transformed = rfecv.transform(X_train_transformed)
                    X_train_transformed = rfecv.fit_transform(X_train_transformed, y_train_o)
                    self.models.append(rfecv)
                    X_test_transformed = rfecv.transform(X_test_transformed)
                    X_minmax = rfecv.transform(X_minmax)
                    features = rfecv.n_features_
                    print("\tShape of transformed train dataset is: {}".format(X_train_transformed.shape))
                    print("\tOptimal no. of features are: {}".format(features))
                    ranking = rfecv.ranking_

                    # Update the feature list here
                    list_temp = self.updateFeatures(list_temp_prev, ranking)

                # This is just used to check the score after inner loop is finished as the test data was already transformed
                # to reduced features. Hence we inverse the transform to check the score
                X_temp = rfecv.inverse_transform(X_test_transformed)
                score = rfecv.score(X_temp, y_test_o)
                self.scores.append(score)
                print("Shape of transformed train dataset is: {}".format(X_train_transformed.shape))
                print("Shape of ranks is: {}\n\n".format(ranking.shape))

        # Print the average scores after finshing the outer loop and save the features in an excel file
        print("After outer loop CV, mean score is: {}".format(mean(self.scores)))
        self.list = list_temp_prev
        self.ranking = ranking
        print(X_train_transformed.shape)
        print(X_test_transformed.shape)
        self.X_transformed = np.vstack((X_train_transformed, X_test_transformed))

        return self

    def predict(self, X):
        X_minmax = self.X_minmax
        y_hat = np.zeros((X.shape[0]))
        for outer in range(self.outer_loop):
            for i, (train_o, test_o) in enumerate(self.loop_indices):
                X_test_o = X_minmax[test_o]
                X_test_transformed = copy.deepcopy(X_test_o)
                for inner in range(self.inner_loop):
                    rfecv = self.models[i]
                    X_test_transformed = rfecv.transform(X_test_transformed)
                    X_minmax = rfecv.transform(X_minmax)

                X_temp = rfecv.inverse_transform(X_test_transformed)
                y_hat[test_o] = rfecv.predict(X_temp)
        return y_hat

    def score(self):
        return self.scores[-1]

    def transformed(self):
        return self.X_transformed
    
    def updateFeatures(self, feature_list, ranks):
        new_list = list()
        for i, r in enumerate(ranks):
            if (r == 1):
                new_list.append(feature_list[i])

        return new_list

    def saveFeatures(self):
        filename = 'Dataset/temp.xlsx'
        wb = load_workbook(filename)
        ws = wb.create_sheet("Final_list")
        col = 1

        for i, r in enumerate(self.ranking):
            if (r == 1):
                wcell1 = ws.cell(col, 1, self.list[i])
                col += 1

        wb.save(filename)