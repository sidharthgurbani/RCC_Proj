from sklearn import svm
from sklearn.model_selection import StratifiedKFold, cross_val_score
from sklearn.ensemble import RandomForestClassifier
from sklearn.feature_selection import RFECV, RFE
from sklearn.preprocessing import StandardScaler, MinMaxScaler, RobustScaler
import matplotlib.pyplot as plt
from openpyxl import *
import copy

def featureSelectAndClassifyRFECV(X_train, X_test, y_train, y_test):

    scaler = MinMaxScaler()
    #scaler = StandardScaler()
    #scaler = RobustScaler()
    X_train_minmax = scaler.fit_transform(X_train)
    X_test_minmax = scaler.transform(X_test)

    #svc =svm.LinearSVC()
    rf = RandomForestClassifier(n_estimators=50, max_depth=20)

    rfecv = RFECV(estimator=rf, step=1, min_features_to_select=5, cv = StratifiedKFold(5), scoring='accuracy')

    X_train_transformed = rfecv.fit_transform(X_train_minmax, y_train)
    #X_train_transformed = rfecv.fit_transform(X_train, y_train)
    X_test_transformed = rfecv.transform(X_test_minmax)
    #X_test_transformed = rfecv.transform(X_test)
    score = rfecv.score(X_test_minmax, y_test)
    #score = rfecv.score(X_test, y_test)

    print('Optimal no. of features are ' + str(rfecv.n_features_))
    print('Score for test set is ' + str(score))
    print(rfecv.ranking_.shape)
    print(X_train_transformed.shape)
    print(X_test_transformed.shape)

    plt.figure()
    plt.xlabel('no. of features')
    plt.ylabel('cv score')
    plt.plot(range(1, len(rfecv.grid_scores_) + 1), rfecv.grid_scores_)
    plt.show()


def featureSelectAndClassifyRFE(X_train, X_test, y_train, y_test):

    scaler = MinMaxScaler()
    X_train_minmax = scaler.fit_transform(X_train)
    #X_test_minmax = scaler.transform(X_test)

    ranks = []
    for i in range(1, 2):
        #svc =svm.LinearSVC()
        rf = RandomForestClassifier(n_estimators=20, max_depth=15)
        rfe = RFE(estimator=rf, step=1, n_features_to_select=300)
        X_train_transformed = rfe.fit_transform(X_train_minmax, y_train)
        print(X_train_transformed.shape)
        #X_test_transformed = rfe.transform(X_test_minmax)

        #score = rfe.score(X_test_minmax, y_test)

        print('Optimal no. of features are ' + str(rfe.n_features_))
        #print('Score for test set is ' + str(score))
        ranks.append(rfe.ranking_)

    print(ranks[0].shape)
    print(ranks[0])
    return ranks[0]


def writeToExcelFile(ranks):
    wb = load_workbook("Dataset/temp.xlsx")
    ws = wb.create_sheet("Sheet2")
    i=5
    for val in ranks:
        wcell = ws.cell(5,i)
        wcell.value = val
        i=i+1

    wb.save("Dataset/temp.xlsx")


# for train_i, test_i in kfold.split(X_train_o, y_train_o):
#     X_train_i = X_train_o[train_i]
#     y_train_i = y_train_o[train_i]
#     X_test_i = X_train_o[test_i]
#     y_test_i = y_train_o[test_i]
#     X_train_transformed = rfe.fit(X_train_i, y_train_i)