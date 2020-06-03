from sklearn import svm
from sklearn.model_selection import StratifiedKFold, cross_val_score
from sklearn.ensemble import RandomForestClassifier
from sklearn.feature_selection import RFECV, RFE
from sklearn.preprocessing import StandardScaler
from openpyxl import load_workbook
from openpyxl.styles import Font
from statistics import mean
import numpy as np
import copy

# This function updates the feature list after every loop to keep track of the features that are reduced.
def updateFeatures(feature_list, ranks):
    new_list = list()
    for i,r in enumerate(ranks):
        if (r==1):
            new_list.append(feature_list[i])

    return new_list

# After nested CV is finished, we want to save the input data with reduced features.
def saveFeatures(feature_list, ranks, X, name, save='feature_names_only'):
    filename = 'Dataset/temp.xlsx'
    wb = load_workbook(filename)
    ws = wb.create_sheet(name)
    col=1
    j=0

    for i,r in enumerate(ranks):
        if (r == 1):
            if save=='feature_names_only':
                wcell1 = ws.cell(col, 1, feature_list[i])
                wcell1.font = Font(bold=True)
                col += 1
            elif save=='feature_names_and_data':
                wcell1 = ws.cell(1, col, feature_list[i])
                wcell1.font = Font(bold=True)
                for row, val in enumerate(X[:,j]):
                    wcell2 = ws.cell(row+2, col ,val)
                j += 1
                col += 1
            else:
                print("\nPlease mention what you want to save!\n")


    wb.save(filename)
    return

# The function performs RFE with CV. This was just used for testing purposes
def eliminateFeaturesRecursively(dataset, X_train, X_test, y_train, y_test, feature_list, clfname):
    if dataset=='pv':
        features = [450, 400, 350, 300, 250, 200, 150, 100, 50]
    elif dataset=='noncon':
        features = [300, 250,200, 150, 100, 50]
    #list_temp = feature_list
    scaler = StandardScaler()
    X_train_minmax = scaler.fit_transform(X_train)
    X_test_minmax = scaler.transform(X_test)
    ranks = []
    scores = []
    for i,feat in enumerate(features):
        if clfname=='svc':
            clf = svm.LinearSVC()
        elif clfname=='rf':
            clf = RandomForestClassifier(n_estimators=50, max_depth=20)

        rfe = RFE(estimator=clf, step=1, n_features_to_select=feat)

        X_train_transformed = rfe.fit_transform(X_train_minmax, y_train)
        X_test_transformed = rfe.transform(X_test_minmax)
        score = rfe.score(X_test_minmax, y_test)
        scores.append(score)
        print("For number of features = {}\n".format(feat))
        print("Shape of transformed train dataset is: {}".format(X_train_transformed.shape))
        print("Optimal no. of features are: {}".format(rfe.n_features_))
        print("Mean Score of transformed dataset with CV is: {:.2f}".format(score))
        ranks.append(rfe.ranking_)
        #name = "Top" + str(feat)
        #list_temp = updateFeatureList(list_temp, rfe.ranking_, X_train_minmax, name)
        print("Shape of ranks is: {}\n\n".format(ranks[i].shape))
        X_train_minmax = copy.deepcopy(X_train_transformed)
        X_test_minmax = copy.deepcopy(X_test_transformed)

    return X_train_minmax, X_test_minmax


# This is the function that performs RFE with CV using nested CV.
def eliminateFeaturesRecursivelyWithCV(X, y, clfname, feature_list):
    # Set the number of inner loops needed to perform. May vary depending on the dataset. Its is suggestive
    # to use atleast 2 for each loop
    outer_loop = 1
    inner_loop = 1

    # Store the original feature list and normalize the data
    list_temp = feature_list
    scaler = StandardScaler()
    X_minmax = scaler.fit_transform(X)
    scores = []

    # Choose which classifier you need to use to perform RFECV with
    if clfname == 'svc':
        clf = svm.LinearSVC()
    elif clfname == 'rf':
        clf = RandomForestClassifier(n_estimators=20, max_depth=10)

    # Determine the number of folds to be used.
    kfold = StratifiedKFold(n_splits=5, shuffle=True)

    for outer in range(outer_loop):
        print("\n--------This is outer loop {}---------\n".format(outer+1))
        # Run the outer loop from here
        for i,(train_o, test_o) in enumerate(kfold.split(X_minmax, y)):
            print("This is set {}".format(i+1))
            X_train_o = X_minmax[train_o]
            y_train_o = y[train_o]
            X_test_o = X_minmax[test_o]
            y_test_o = y[test_o]
            X_train_transformed = copy.deepcopy(X_train_o)
            X_test_transformed = copy.deepcopy(X_test_o)

            # Run the inner loop from here
            for inner in range(inner_loop):
                # If the number of features are very high (>100), we set the minimum number of features needed to be 100.
                # If the numnber of features are moderate (15-100), we set the minimum number of features to be 10
                # less than already present
                n_feat = min(100, X_train_transformed.shape[1]-10)

                # If the number of features are less (<15), then we want it to select atleast 5 features to continue the loop
                n_feat = max(5, n_feat)
                list_temp_prev = list_temp
                print("\n\t--------This is inner loop {}---------\n".format(inner+1))
                rfecv = RFECV(estimator=clf, step=1, min_features_to_select=n_feat, cv=kfold, scoring='accuracy')

                # Transform the datasets at each loop to keep track of reduced features
                X_train_transformed = rfecv.fit_transform(X_train_transformed, y_train_o)
                X_test_transformed = rfecv.transform(X_test_transformed)
                X_minmax = rfecv.transform(X_minmax)
                features = rfecv.n_features_
                print("\tShape of transformed train dataset is: {}".format(X_train_transformed.shape))
                print("\tOptimal no. of features are: {}".format(features))
                ranking = rfecv.ranking_

                # Update the feature list here
                list_temp = updateFeatures(list_temp_prev, ranking)

            # This is just used to check the score after inner loop is finished as the test data was already transformed
            # to reduced features. Hence we inverse the transform to check the score
            X_temp = rfecv.inverse_transform(X_test_transformed)
            score = rfecv.score(X_temp, y_test_o)
            scores.append(score)
            print("Shape of transformed train dataset is: {}".format(X_train_transformed.shape))
            print("Shape of ranks is: {}\n\n".format(ranking.shape))

    # Print the average scores after finshing the outer loop and save the features in an excel file
    print("After outer loop CV, mean score is: {}".format(mean(scores)))
    X_final = np.vstack((X_train_transformed, X_test_transformed))
    saveFeatures(list_temp_prev, ranking, X_final, 'Final_List')

    return X_final

# This runs SVM and RF forest on the dataset and gives a CV score average.
def noRFE(X, y, clfname, scale=False):
    if scale:
        scaler = StandardScaler()
        X = scaler.fit_transform(X)
    if clfname=='svc':
        clf = svm.LinearSVC()
    elif clfname=='rf':
        clf = RandomForestClassifier(n_estimators=10, max_depth=20)

    score = cross_val_score(clf, X, y, cv=StratifiedKFold(n_splits=5, shuffle=True))
    print("{} Classifier gives mean accuracy after CV {:.2f}%".format(clfname, score.mean() * 100))

def getFeatureWeights(feature_list, features):
    filename = 'Dataset/temp.xlsx'
    wb = load_workbook(filename)
    ws = wb.create_sheet('feature_weights')
    wcell1 = ws.cell(1,1, 'Feature')
    wcell1.font = Font(bold=True)
    wcell2 = ws.cell(1,2, 'Weights')
    wcell2.font = Font(bold=True)
    index = 0
    for i,val in enumerate(features):
        if val > 0.01:
            wcell1 = ws.cell(index+2, 1, feature_list[i])
            wcell2 = ws.cell(index+2, 2, val)
            index += 1

    wb.save(filename)
    return

def updateList(X, feat_imp):
    #new_feat = np.zeros((1))
    new_feat = list()
    #index = 1
    X_new = np.zeros((X.shape[0]))
    for i, val in enumerate(feat_imp):
        if val > 0.01:
            new_feat.append(val)
            col = X[:,i]
            X_new = np.vstack((X_new, col))

    X_new = X_new[1:,:].T
    return X_new, new_feat


def justRF(X, y, feature_list):
    scaler = StandardScaler()
    X = scaler.fit_transform(X)
    clf = RandomForestClassifier(n_estimators=10, max_depth=20)
    clf.fit(X,y)
    feat_imp = clf.feature_importances_
    print(feat_imp.shape)
    print(feat_imp)
    getFeatureWeights(feature_list, feat_imp)
    X_new, new_feat = updateList(X, feat_imp)
    print(X_new.shape)
    print(len(new_feat))
    return X_new

# def updateFeatureList(feature_list, ranks, X, name):
#     wb = load_workbook("Dataset/temp.xlsx")
#     ws = wb.create_sheet(name)
#     new_list = list()
#     filename = 'Dataset/temp.xlsx'
#     #wcell1 = ws.cell(1, 1, "Feature")
#     #wcell2 = ws.cell(1, 2, "Rank")
#     col=1
#     #dict = {}
#     for i,r in enumerate(ranks):
#         if (r == 1):
#             new_list.append(feature_list[i])
#             wcell1 = ws.cell(1, col, feature_list[i])
#             for row, val in enumerate(X[:,i]):
#                 wcell2 = ws.cell(row+2, col ,val)
#             col += 1
#
#     wb.save(filename)
#     return new_list
